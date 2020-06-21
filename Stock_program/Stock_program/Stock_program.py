import requests
import openpyxl
import string
import pandas as pd 
from openpyxl import load_workbook
from bs4 import BeautifulSoup
 
load_wb = load_workbook("C:/list.xlsx", data_only=True)
load_ws = load_wb['List']

wb = openpyxl.Workbook()
sheet = wb.active 
sheet.title = 'Stock'

# 종목 코드
company_codes = []
# 종목 이름
company_names = []

for row in range(2,2207):
    #print(load_ws.cell(row,1).value)
    company_codes.append(load_ws.cell(row,2).value)

for row in range(2,2207):
    #print(load_ws.cell(row,2).value)
    company_names.append(load_ws.cell(row,1).value)

# 아이투자 사이트
def get_iinvest_info(company_code):
    iinvest_df = pd.read_html("http://search.itooza.com/search.htm?seName=" + company_code)
    return iinvest_df

def get_eps(company_code):
    #print(company_code)
    iinvest_df = get_iinvest_info(company_code)
    #print(iinvest_df)
    #연간자료
    coinfo_year = iinvest_df[3]
    #print(coinfo_year)

    #분기자료
    coinfo_quater = iinvest_df[4]
    #print(coinfo_quater)

    #최근 3분기의 EPS가 오르고 있을 때,
    last_quater_three_eps = coinfo_quater.values[0][1:8]

    first_quater_eps = last_quater_three_eps[0]
    second_quater_eps = last_quater_three_eps[1]
    third_quater_eps = last_quater_three_eps[2]

    forth_quater_eps = last_quater_three_eps[3]
    fifth_quater_eps = last_quater_three_eps[4]
    sixth_quater_eps = last_quater_three_eps[5]
    seventh_quater_eps = last_quater_three_eps[6]
    
    
    #print(first_quater_eps,second_quater_eps,third_quater_eps)
   
    if(fifth_quater_eps < first_quater_eps):
        if(sixth_quater_eps < second_quater_eps):
            if(seventh_quater_eps < third_quater_eps):
                quater_flag = 1
            else:
                quater_flag = 0
        else:
            quater_flag = 0
    else:
        quater_flag = 0

    #작년 EPS가 내려가고 있을 때
    last_year_two_eps = coinfo_year.values[0][2:4]

    first_year_eps = last_year_two_eps[0]
    second_year_eps = last_year_two_eps[1]

    #print(first_year_eps,second_year_eps)

    if(second_year_eps > first_year_eps):
        year_flag = 1
    else:
        year_flag = 0

    flag = quater_flag * year_flag

    return flag

# 네이버금융
def get_naver_info(company_code):
    url = "https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd=" + company_code
    result = requests.get(url)
    naver_df = BeautifulSoup(result.content, "html.parser")
    return naver_df

def get_index(company_code):
    coinfo_link1 = get_naver_info(company_code)
    table = coinfo_link1.find("td", {"class": "cmp-table-cell td0301"})

    now_per = table.select('dt')[2].b.get_text()
    if now_per == "":
        now_per = "0"
    now_per = float(now_per.replace(",",""))

    now_avgper = table.select('dt')[3].b.get_text()
    if now_avgper == "":
        now_avgper = "0"
    now_avgper = float(now_avgper.replace(",",""))

    now_pbr = table.select('dt')[4].b.get_text()
    if now_pbr == "":
        now_pbr = "0"
    now_pbr = float(now_pbr.replace(",",""))

    if(now_avgper > now_per):
        flag = 1
    else:
        flag = 0
        
    return flag

for index, item in enumerate(company_codes):
    flag_eps = get_eps(item)
    flag_index = get_index(item)
    flag = flag_index * flag_eps

    if(flag == 1):
        print('%10s'%company_names[index],'%10s'%company_codes[index])

